import yaml
import urllib3
from os.path import expanduser, pathsep, dirname, realpath
import sys

from openpyxl import Workbook

from jazz_dng_client import *
from jira_class import Jira, get_query
import utility_funcs.logger_yaml as log


JIRA_CONFIG_PATH = f"{dirname(realpath(sys.argv[0]))}/config.yaml{pathsep}~/.jira/config.yaml"


def read_jazz_requirements(jazz_client, path: str, name: str) -> dict:
    requirements_by_id = {}
    log.logger.info(f"Opening the collection {name}")
    result_list = Folder(jazz_client).get_folder_artifacts(path=path, name=name)
    if len(result_list) < 1 :
        log.logger.error(f"Unable to open collection, path='{path}', name='{name}'; not found or possibly locked.")
        exit(-1)

    for resource_collection in jazz.get_object_from_uri(result_list[0]):
        if isinstance(resource_collection, Collection):
            requirement_set = resource_collection.requirement_set()
            log.logger.info(f"The collection contains {len(requirement_set)} requirements, reading...")
            requirements_by_id = { requirement.get_identifier(): requirement for requirement in jazz.get_object_from_uri(requirement_set)}
            log.logger.info(f"got {len(requirements_by_id)} Requirements")
        else:
            log.logger.error(f"Object {resource_collection} is not a collection")
            exit(-1)
    return requirements_by_id


def read_jazz(jazz_client):
    return read_jazz_requirements(jazz_client=jazz_client,
                                  path=environment['path'],
                                  name=environment['name'])


def read_jira_epics(jira_client, query):
    epic_name = jira_client.get_field_name('Epic Name')
    items_by_id = {getattr(item.fields, epic_name): item
                   for item in jira_client.do_query(query)}
    return items_by_id


def read_jira(jira_client):
    return read_jira_epics(jira_client=jira_client, query=environment['jira_query'])


def sort_matched(y):
    requirement, item = y
    return requirement.get_identifier()


def set_width(sheet, cols: dict):
    for col in cols:
        sheet.column_dimensions[col].width = cols[col]


try:
    with open("environments.yaml", "r") as f:
        config = yaml.load(f)
except FileNotFoundError as nf:
    log.logger.error("File 'environments.yaml' not found %s", nf)
    exit(-1)

if len(sys.argv) < 2:
    log.logger.error(f"{sys.argv[0]} <tag>, available tags are {[t for t in config]}" )
    exit(-1)

if sys.argv[1] not in config:
    log.logger.error(f"configuration '{sys.argv[1]}' was not found")
    exit(-1)

environment = config[sys.argv[1]]

urllib3.disable_warnings()
jazz = Jazz(server_alias=environment['jazz_server'], config_path=JAZZ_CONFIG_PATH, use_cache=True, op_name=None)
jira = Jira(environment['jira_server'], JIRA_CONFIG_PATH, log=log.logger)
epic_name = jira.get_field_name('Epic Name')
status = jira.get_field_name('Status')
external_link = jira.get_field_name('External Link')
requirements_by_id = read_jazz(jazz_client=jazz)
items_by_id = read_jira(jira_client=jira)

log.logger.info("Analyzing...")
missing_requirements = [requirement for key, requirement in requirements_by_id.items()
                        if requirement.get_identifier() not in items_by_id]
missing_requirements = sorted(missing_requirements, key=lambda y: y.get_identifier())
missing_items = [item for key, item in items_by_id.items()
                 if getattr(item.fields, epic_name) not in requirements_by_id]
missing_items = sorted(missing_items, key=lambda item: getattr(item.fields, epic_name))
matched = [(requirement, items_by_id[key]) for key, requirement in requirements_by_id.items()
           if requirement.get_identifier() in items_by_id]

matched = sorted(matched, key=sort_matched)

wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Summary"
ws_summary.column_dimensions['B'].width = 12

ws_requirements = wb.create_sheet(title="Requirements not in Jira")
ws_items = wb.create_sheet(title="Jira items not in Collection")
ws_matching = wb.create_sheet(title="Matched DNG and Jira Items")

log.logger.info("--------------------------------------------------------------------------------")
log.logger.info(f"Summary:")
ws_summary['A1'] = "Summary:"
log.logger.info(f"{len(requirements_by_id)} DNG Requirements were analyzed")
ws_summary['B3'] = "DNG Path:"
ws_summary['C3'] = environment['path']
ws_summary['B4'] = "DNG Name:"
ws_summary['C4'] = environment['name']
ws_summary['B5'] = "Jira Query:"
ws_summary['C5'] = environment['jira_query']
ws_summary['B6'] = len(requirements_by_id)
ws_summary['C6'] = "DNG Requirements were analyzed"
log.logger.info(f"{len(items_by_id)} Jira items were analyzed")
ws_summary['B7'] = len(items_by_id)
ws_summary['C7'] = "Jira items were analyzed"
log.logger.info(f"{len(missing_requirements)} Requirements in collection, not found in Jira")
ws_summary['B8'] = len(missing_requirements)
ws_summary['C8'] = "Requirements in collection, not found in Jira"
log.logger.info(f"{len(missing_items)} Jira items not in specified Requirement Collection")
ws_summary['B9'] = len(missing_items)
ws_summary['C9'] = "Jira items not in specified Requirement Collection"
log.logger.info(f"{len(matched)} DNG Requirements and Jira Items are matched")
ws_summary['B10'] = len(matched)
ws_summary['C10'] = "DNG Requirements and Jira Items are matched"
log.logger.info("--------------------------------------------------------------------------------")
ws_requirements['A1'] = f"Requirements from Collection '{environment['name']}'not found in Jira:"
set_width(ws_requirements, {'A': 12, 'B': 40})
ws_requirements.append(['-'])
ws_requirements.append(['DNG ID', 'DNG Name'])
for requirement in missing_requirements:
    ws_requirements.append([requirement.get_identifier(), requirement.get_name()])

ws_items['A1'] = f"Jira items without matching DNG Requirement in Collection '{environment['name']}':"
set_width(ws_items, {'A': 12, 'B': 8, 'C': 9, 'D': 80, 'E': 80})
ws_items.append(['-'])
ws_items.append(['Jira Key', 'Epic', 'Status', 'Summary', 'Link to DNG'])
for item in missing_items:
    ws_items.append([item.key,
                     '=HYPERLINK("'+getattr(item.fields, external_link)+'","'+str(getattr(item.fields, epic_name))+'")',
                     str(getattr(item.fields, status)),
                     item.fields.summary,
                     '=HYPERLINK("'+getattr(item.fields, external_link)+'")'])

ws_matching['A1'] = f"DNG Requirements and Jira Items that match from Collection '{environment['name']}':"
set_width(ws_matching, {'A': 10, 'B': 10, 'C': 8, 'D': 9, 'E': 80, 'F': 80})
ws_matching.append(['-'])
ws_matching.append(['DNG ID', 'Jira Key', 'Epic', 'Status', 'Summary', 'Link to DNG'])
for x in matched:
    requirement, item = x
    ws_matching.append(['='+requirement.get_identifier(),
                        item.key,
                        '=HYPERLINK("'+getattr(item.fields, external_link)+'","'+getattr(item.fields, epic_name)+'")',
                        str(getattr(item.fields, status)), item.fields.summary,
                        '=HYPERLINK("'+getattr(item.fields, external_link)+'")'])

wb.save(environment['xls'])
log.logger.info("Done.")



















